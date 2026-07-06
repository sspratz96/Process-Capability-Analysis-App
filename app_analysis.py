import math
import multiprocessing as mp
import queue
from dataclasses import dataclass

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
import statsmodels.api as sm
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from scipy import stats


GOOD_FIT_P_VALUE = 0.05
RANDOM_SEED = 42
DEFAULT_FIT_TIMEOUT_SECONDS = 8


@dataclass
class FitResult:
    name: str
    display_name: str
    params: tuple
    ks_statistic: float
    p_value: float
    aic: float
    bic: float


DIST_CANDIDATES = [
    "alpha",
    "anglit",
    "arcsine",
    "argus",
    "beta",
    "betaprime",
    "bradford",
    "burr",
    "burr12",
    "cauchy",
    "chi",
    "chi2",
    "cosine",
    "crystalball",
    "dgamma",
    "dweibull",
    "erlang",
    "expon",
    "exponnorm",
    "exponpow",
    "exponweib",
    "f",
    "fatiguelife",
    "fisk",
    "foldcauchy",
    "foldnorm",
    "gamma",
    "gausshyper",
    "genexpon",
    "genextreme",
    "gengamma",
    "genhalflogistic",
    "genhyperbolic",
    "geninvgauss",
    "genlogistic",
    "gennorm",
    "genpareto",
    "gibrat",
    "gompertz",
    "gumbel_l",
    "gumbel_r",
    "halfcauchy",
    "halfgennorm",
    "halflogistic",
    "halfnorm",
    "hypsecant",
    "invgamma",
    "invgauss",
    "invweibull",
    "johnsonsb",
    "johnsonsu",
    "kappa3",
    "kappa4",
    "ksone",
    "kstwo",
    "kstwobign",
    "laplace",
    "laplace_asymmetric",
    "levy",
    "levy_l",
    "loggamma",
    "logistic",
    "loglaplace",
    "lognorm",
    "lomax",
    "maxwell",
    "mielke",
    "moyal",
    "nakagami",
    "ncf",
    "nct",
    "ncx2",
    "norm",
    "norminvgauss",
    "pareto",
    "pearson3",
    "powerlaw",
    "powerlognorm",
    "powernorm",
    "rayleigh",
    "rdist",
    "recipinvgauss",
    "reciprocal",
    "rice",
    "semicircular",
    "skewcauchy",
    "skewnorm",
    "studentized_range",
    "t",
    "trapezoid",
    "trapz",
    "triang",
    "truncexpon",
    "truncnorm",
    "tukeylambda",
    "uniform",
    "vonmises",
    "vonmises_line",
    "wald",
    "weibull_max",
    "weibull_min",
    "wrapcauchy",
]


def init_state():
    defaults = {
        "data": None,
        "file_name": None,
        "selected_col": None,
        "normal_result": None,
        "normal_trim_percent": 0.0,
        "normal_applied_trim_percent": 0.0,
        "nonnormal_trim_percent": 0.0,
        "nonnormal_applied_trim_percent": 0.0,
        "selected_fit_name": None,
        "monte_carlo_result": None,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def clear_analysis_state():
    st.session_state.selected_col = None
    st.session_state.normal_result = None
    st.session_state.normal_trim_percent = 0.0
    st.session_state.normal_applied_trim_percent = 0.0
    st.session_state.nonnormal_result = None
    st.session_state.nonnormal_trim_percent = 0.0
    st.session_state.nonnormal_applied_trim_percent = 0.0
    st.session_state.selected_fit_name = None
    st.session_state.monte_carlo_result = None


def read_csv_file(uploaded_file):
    return pd.read_csv(uploaded_file)


def read_strict_xlsx_table(uploaded_file):
    workbook = load_workbook(uploaded_file, data_only=True)
    if len(workbook.sheetnames) != 1:
        raise ValueError("El archivo .xlsx debe tener exactamente una hoja.")

    worksheet = workbook[workbook.sheetnames[0]]
    tables = list(worksheet.tables.values())
    if len(tables) != 1:
        raise ValueError("La unica hoja debe contener exactamente una tabla de Excel.")

    table = tables[0]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            inside_table = min_row <= cell.row <= max_row and min_col <= cell.column <= max_col
            if not inside_table:
                raise ValueError("No debe haber datos fuera de la tabla de Excel.")

    rows = []
    for row in worksheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    ):
        rows.append(list(row))

    if len(rows) < 2:
        raise ValueError("La tabla debe tener encabezados y al menos una fila de datos.")

    headers = [str(header).strip() if header is not None else "" for header in rows[0]]
    if any(not header for header in headers):
        raise ValueError("La tabla no puede tener encabezados vacios.")
    if len(set(headers)) != len(headers):
        raise ValueError("La tabla no puede tener encabezados repetidos.")

    df = pd.DataFrame(rows[1:], columns=headers)
    df = df.dropna(how="all")
    return df


def load_uploaded_file(uploaded_file):
    file_name = uploaded_file.name
    lower_name = file_name.lower()
    if lower_name.endswith(".csv"):
        return read_csv_file(uploaded_file)
    if lower_name.endswith(".xlsx"):
        return read_strict_xlsx_table(uploaded_file)
    if lower_name.endswith(".xlsm"):
        raise ValueError("Los archivos .xlsm no estan soportados. Guarda una copia como .xlsx sin macros o exporta la tabla como .csv.")
    raise ValueError("Formato no soportado. Usa un archivo .csv o .xlsx.")


def numeric_columns(df):
    numeric = []
    for column in df.columns:
        series = df[column]
        if pd.api.types.is_datetime64_any_dtype(series):
            continue
        converted = pd.to_numeric(series, errors="coerce")
        if converted.notna().sum() > 0 and converted.notna().sum() == series.notna().sum():
            numeric.append(column)
    return numeric


def clean_numeric_series(df, column):
    series = pd.to_numeric(df[column], errors="coerce").dropna()
    series = series.replace([np.inf, -np.inf], np.nan).dropna()
    return series.astype(float)


def trim_outliers(series, percent):
    if percent <= 0:
        return series.copy()
    tail = min(percent / 200, 0.49)
    lower = series.quantile(tail)
    upper = series.quantile(1 - tail)
    return series[(series >= lower) & (series <= upper)].copy()


def normality_test(series):
    if len(series) < 8:
        return np.nan, np.nan, False
    mean = series.mean()
    std = series.std(ddof=1)
    if std <= 0 or not np.isfinite(std):
        return np.nan, np.nan, False
    fitted_normal = stats.norm(loc=mean, scale=std)
    _, p_value = stats.kstest(series.to_numpy(), fitted_normal.cdf)
    return mean, p_value, p_value > GOOD_FIT_P_VALUE


def process_judgement(index_value):
    if index_value is None or not np.isfinite(index_value):
        return "No calculable"
    if index_value < 1:
        return "Se debe mejorar"
    if index_value <= 2:
        return "Susceptible a mejora"
    return "Proceso capaz"


def normal_capability(series, il, ul):
    mean = series.mean()
    sigma = series.std(ddof=1)
    if sigma <= 0 or not np.isfinite(sigma):
        return {"mean": mean, "sigma": sigma, "cp": np.nan, "cpk": np.nan, "judgement": "No calculable"}

    has_il = il is not None and np.isfinite(il)
    has_ul = ul is not None and np.isfinite(ul)
    cp = np.nan
    lower_index = np.nan
    upper_index = np.nan

    if has_il:
        lower_index = (mean - il) / (3 * sigma)
    if has_ul:
        upper_index = (ul - mean) / (3 * sigma)
    if has_il and has_ul:
        cp = (ul - il) / (6 * sigma)
        cpk = min(lower_index, upper_index)
    elif has_il:
        cpk = lower_index
    elif has_ul:
        cpk = upper_index
    else:
        cpk = np.nan

    return {"mean": mean, "sigma": sigma, "cp": cp, "cpk": cpk, "judgement": process_judgement(cpk)}


def nonnormal_capability(series, distribution, params, il, ul):
    median = distribution.ppf(0.5, *params)
    q_low = distribution.ppf(0.00135, *params)
    q_high = distribution.ppf(0.99865, *params)
    has_il = il is not None and np.isfinite(il)
    has_ul = ul is not None and np.isfinite(ul)

    lower_index = np.nan
    upper_index = np.nan
    if has_il and np.isfinite(q_low) and median != q_low:
        lower_index = (median - il) / (median - q_low)
    if has_ul and np.isfinite(q_high) and q_high != median:
        upper_index = (ul - median) / (q_high - median)

    if has_il and has_ul:
        ppk = min(lower_index, upper_index)
    elif has_il:
        ppk = lower_index
    elif has_ul:
        ppk = upper_index
    else:
        ppk = np.nan

    return {"median": median, "q_low": q_low, "q_high": q_high, "ppk": ppk, "judgement": process_judgement(ppk)}


def specification_inputs(prefix):
    col1, col2 = st.columns(2)
    with col1:
        use_il = st.checkbox("Usar IL / LSL", value=True, key=f"{prefix}_use_il")
        il = st.number_input("IL / LSL", value=0.0, key=f"{prefix}_il") if use_il else None
    with col2:
        use_ul = st.checkbox("Usar UL / USL", value=True, key=f"{prefix}_use_ul")
        ul = st.number_input("UL / USL", value=1.0, key=f"{prefix}_ul") if use_ul else None
    if not use_il and not use_ul:
        st.warning("Debes definir al menos IL o UL para calcular capacidad.")
    return il, ul


def draw_reference_lines(ax, il, ul):
    if il is not None and np.isfinite(il):
        ax.axvline(il, color="#7c2d12", linestyle="--", linewidth=1.5, label="IL")
    if ul is not None and np.isfinite(ul):
        ax.axvline(ul, color="#1e3a8a", linestyle="--", linewidth=1.5, label="UL")
    if ax.get_legend_handles_labels()[0]:
        ax.legend()


def plot_normal_analysis(series, simulated, is_normal, il, ul, title_suffix):
    color = "#15803d" if is_normal else "#b91c1c"
    fig = plt.figure(figsize=(14, 9))

    ax1 = fig.add_subplot(2, 2, 1)
    sm.qqplot(series, line="s", ax=ax1)
    ax1.set_title(f"Q-Q Plot datos reales {title_suffix}")

    ax2 = fig.add_subplot(2, 2, 2)
    sns.kdeplot(series, ax=ax2, color="#2563eb", label="Datos reales", linewidth=2)
    sns.kdeplot(simulated, ax=ax2, color=color, label="Normal ideal simulada", linewidth=2, linestyle="--")
    draw_reference_lines(ax2, il, ul)
    ax2.set_title("Density Plot")
    ax2.legend()

    ax3 = fig.add_subplot(2, 2, 3)
    plot_data = pd.DataFrame({"Datos reales": series.reset_index(drop=True), "Normal ideal": simulated})
    sns.boxplot(data=plot_data, ax=ax3, palette=["#93c5fd", color])
    ax3.set_title("Box Plot")

    ax4 = fig.add_subplot(2, 2, 4)
    ax4.plot(series.reset_index(drop=True), color="#2563eb", label="Datos reales")
    ax4.plot(simulated, color=color, alpha=0.8, label="Normal ideal simulada")
    if il is not None and np.isfinite(il):
        ax4.axhline(il, color="#7c2d12", linestyle="--", linewidth=1.5, label="IL")
    if ul is not None and np.isfinite(ul):
        ax4.axhline(ul, color="#1e3a8a", linestyle="--", linewidth=1.5, label="UL")
    ax4.set_title("Serie de Tiempo")
    ax4.set_xlabel("Indice")
    ax4.legend()

    fig.tight_layout()
    st.pyplot(fig)


def plot_distribution_analysis(series, distribution, params, is_good_fit, il, ul, title_suffix):
    color = "#15803d" if is_good_fit else "#b91c1c"
    x_min, x_max = series.min(), series.max()
    padding = (x_max - x_min) * 0.15 if x_max > x_min else 1
    x = np.linspace(x_min - padding, x_max + padding, 500)
    pdf = distribution.pdf(x, *params)
    simulated = distribution.rvs(*params, size=len(series), random_state=RANDOM_SEED)

    fig = plt.figure(figsize=(14, 9))

    ax1 = fig.add_subplot(2, 2, 1)
    sorted_real = np.sort(series)
    theoretical = distribution.ppf((np.arange(1, len(series) + 1) - 0.5) / len(series), *params)
    ax1.scatter(theoretical, sorted_real, color="#2563eb", alpha=0.75)
    line_min = min(np.nanmin(theoretical), np.nanmin(sorted_real))
    line_max = max(np.nanmax(theoretical), np.nanmax(sorted_real))
    ax1.plot([line_min, line_max], [line_min, line_max], color=color, linestyle="--")
    ax1.set_title(f"Q-Q Plot fit {title_suffix}")

    ax2 = fig.add_subplot(2, 2, 2)
    sns.kdeplot(series, ax=ax2, color="#2563eb", label="Datos reales", linewidth=2)
    ax2.plot(x, pdf, color=color, label="Distribucion ajustada", linewidth=2, linestyle="--")
    draw_reference_lines(ax2, il, ul)
    ax2.set_title("Density Plot")
    ax2.legend()

    ax3 = fig.add_subplot(2, 2, 3)
    plot_data = pd.DataFrame({"Datos reales": series.reset_index(drop=True), "Fit simulado": simulated})
    sns.boxplot(data=plot_data, ax=ax3, palette=["#93c5fd", color])
    ax3.set_title("Box Plot")

    ax4 = fig.add_subplot(2, 2, 4)
    ax4.plot(series.reset_index(drop=True), color="#2563eb", label="Datos reales")
    ax4.plot(simulated, color=color, alpha=0.8, label="Fit simulado")
    if il is not None and np.isfinite(il):
        ax4.axhline(il, color="#7c2d12", linestyle="--", linewidth=1.5, label="IL")
    if ul is not None and np.isfinite(ul):
        ax4.axhline(ul, color="#1e3a8a", linestyle="--", linewidth=1.5, label="UL")
    ax4.set_title("Serie de Tiempo")
    ax4.set_xlabel("Indice")
    ax4.legend()

    fig.tight_layout()
    st.pyplot(fig)


def fit_distribution_values(values, name):
    distribution = getattr(stats, name)
    params = distribution.fit(values)
    fitted_distribution = distribution(*params)
    ks_statistic, p_value = stats.kstest(values, fitted_distribution.cdf)
    log_pdf = fitted_distribution.logpdf(values)
    log_pdf = log_pdf[np.isfinite(log_pdf)]
    if len(log_pdf) == 0:
        raise ValueError("Log-likelihood no finito")
    log_likelihood = np.sum(log_pdf)
    param_count = len(params)
    aic = 2 * param_count - 2 * log_likelihood
    bic = math.log(len(values)) * param_count - 2 * log_likelihood
    return FitResult(
        name=name,
        display_name=name.replace("_", " ").title(),
        params=params,
        ks_statistic=ks_statistic,
        p_value=p_value,
        aic=aic,
        bic=bic,
    )


def fit_distribution(series, name):
    return fit_distribution_values(series.to_numpy(), name)


def fit_distribution_worker_loop(values, task_queue, output_queue):
    while True:
        task = task_queue.get()
        if task is None:
            break
        task_id, name = task
        try:
            result = fit_distribution_values(values, name)
            output_queue.put((task_id, "ok", result))
        except Exception as exc:
            output_queue.put((task_id, "error", str(exc)))


def start_fit_worker(values):
    context = mp.get_context("spawn")
    task_queue = context.Queue()
    output_queue = context.Queue()
    process = context.Process(target=fit_distribution_worker_loop, args=(values, task_queue, output_queue))
    process.start()
    return process, task_queue, output_queue


def stop_fit_worker(process, task_queue, output_queue):
    if process.is_alive():
        try:
            task_queue.put(None)
            process.join(1)
        except Exception:
            pass
    if process.is_alive():
        process.terminate()
        process.join(1)
    task_queue.close()
    output_queue.close()


def fit_many_distributions(series, timeout_seconds=DEFAULT_FIT_TIMEOUT_SECONDS):
    results = []
    skipped_by_timeout = []
    values = series.to_numpy()
    process, task_queue, output_queue = start_fit_worker(values)
    progress = st.progress(0, text="Probando distribuciones...")
    try:
        for index, name in enumerate(DIST_CANDIDATES[:100], start=1):
            task_queue.put((index, name))
            try:
                task_id, status, payload = output_queue.get(timeout=timeout_seconds)
                if task_id == index and status == "ok" and np.isfinite(payload.aic) and np.isfinite(payload.p_value):
                    results.append(payload)
            except queue.Empty:
                skipped_by_timeout.append(name)
                stop_fit_worker(process, task_queue, output_queue)
                process, task_queue, output_queue = start_fit_worker(values)
            except Exception:
                pass
            progress.progress(
                index / min(100, len(DIST_CANDIDATES)),
                text=f"Probando distribuciones... {index}/100 | Saltadas por timeout: {len(skipped_by_timeout)}",
            )
    finally:
        stop_fit_worker(process, task_queue, output_queue)
        progress.empty()
    return sorted(results, key=lambda item: (item.aic, -item.p_value))[:10], skipped_by_timeout


def results_to_dataframe(results):
    return pd.DataFrame(
        [
            {
                "Distribucion": item.display_name,
                "KS statistic": item.ks_statistic,
                "p-value": item.p_value,
                "Buen fit": "Si" if item.p_value > GOOD_FIT_P_VALUE else "No",
                "AIC": item.aic,
                "BIC": item.bic,
            }
            for item in results
        ]
    )


def simulate_normal_for_target(series, il, ul, target_index, n_simulations):
    mean = series.mean()
    if il is not None and ul is not None:
        sigma = (ul - il) / (6 * target_index)
        mean = (il + ul) / 2
    elif il is not None:
        sigma = max((mean - il) / (3 * target_index), series.std(ddof=1) / target_index)
    elif ul is not None:
        sigma = max((ul - mean) / (3 * target_index), series.std(ddof=1) / target_index)
    else:
        sigma = series.std(ddof=1)
    return stats.norm.rvs(loc=mean, scale=abs(sigma), size=n_simulations, random_state=RANDOM_SEED)


def run_monte_carlo(source_label, distribution, params, real_series, il, ul, target_index, n_simulations):
    real_simulated = distribution.rvs(*params, size=n_simulations, random_state=RANDOM_SEED)

    if distribution.name == "norm":
        target_simulated = simulate_normal_for_target(real_series, il, ul, target_index, n_simulations)
    else:
        current_capability = nonnormal_capability(real_series, distribution, params, il, ul)
        current_index = current_capability["ppk"]
        if current_index and np.isfinite(current_index) and current_index > 0:
            median = np.nanmedian(real_simulated)
            target_simulated = median + (real_simulated - median) * (current_index / target_index)
        else:
            target_simulated = real_simulated.copy()

    fig = plt.figure(figsize=(12, 5))
    ax = fig.add_subplot(1, 1, 1)
    sns.kdeplot(real_simulated, ax=ax, color="#2563eb", label=f"Simulacion real: {source_label}", linewidth=2)
    sns.kdeplot(target_simulated, ax=ax, color="#15803d", label="Simulacion objetivo", linewidth=2, linestyle="--")
    draw_reference_lines(ax, il, ul)
    ax.set_title("Monte Carlo: capacidad real vs capacidad objetivo")
    ax.legend()
    st.pyplot(fig)

    return {
        "real_mean": float(np.mean(real_simulated)),
        "target_mean": float(np.mean(target_simulated)),
        "real_std": float(np.std(real_simulated, ddof=1)),
        "target_std": float(np.std(target_simulated, ddof=1)),
    }


def normal_analysis_tab(series):
    st.subheader("Analisis normal")
    trim_percent = st.slider(
        "Quitar outliers (%)",
        min_value=0.0,
        max_value=30.0,
        value=float(st.session_state.normal_trim_percent),
        step=0.5,
        key="normal_trim_slider",
    )
    il, ul = specification_inputs("normal")

    if st.button("Aplicar analisis normal", type="primary"):
        trimmed = trim_outliers(series, trim_percent)
        mean, p_value, is_normal = normality_test(trimmed)
        simulated = stats.norm.rvs(
            loc=trimmed.mean(),
            scale=trimmed.std(ddof=1),
            size=len(trimmed),
            random_state=RANDOM_SEED,
        )
        st.session_state.normal_result = {
            "series": trimmed,
            "p_value": p_value,
            "is_normal": is_normal,
            "simulated": simulated,
            "capability": normal_capability(trimmed, il, ul),
            "il": il,
            "ul": ul,
        }
        st.session_state.normal_applied_trim_percent = trim_percent

    result = st.session_state.normal_result
    if result:
        capability = result["capability"]
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Normalidad", "Normal" if result["is_normal"] else "No normal")
        col2.metric("p-value KS", f"{result['p_value']:.4f}" if np.isfinite(result["p_value"]) else "N/A")
        col3.metric("Cpk", f"{capability['cpk']:.3f}" if np.isfinite(capability["cpk"]) else "N/A")
        col4.metric("Estado", capability["judgement"])

        with st.expander("Estadisticas descriptivas", expanded=True):
            st.write(result["series"].describe())
            st.write(
                {
                    "Media": capability["mean"],
                    "Sigma": capability["sigma"],
                    "Cp": capability["cp"],
                    "Cpk": capability["cpk"],
                    "Outliers removidos (%)": st.session_state.normal_applied_trim_percent,
                }
            )

        plot_normal_analysis(
            result["series"],
            result["simulated"],
            result["is_normal"],
            result["il"],
            result["ul"],
            "",
        )

        if not result["is_normal"]:
            st.warning("Recomendamos hacer un analisis no normal. No es excluyente: puedes hacerlo aunque el test normal sea aceptable.")


def nonnormal_analysis_tab(series):
    st.subheader("Analisis no normal")
    col_trim, col_timeout = st.columns(2)
    with col_trim:
        trim_percent = st.slider(
            "Quitar outliers (%)",
            min_value=0.0,
            max_value=30.0,
            value=float(st.session_state.nonnormal_trim_percent),
            step=0.5,
            key="nonnormal_trim_slider",
        )
    with col_timeout:
        timeout_seconds = st.number_input(
            "Tiempo maximo por distribucion (segundos)",
            min_value=1,
            max_value=60,
            value=DEFAULT_FIT_TIMEOUT_SECONDS,
            step=1,
        )
    il, ul = specification_inputs("nonnormal")

    if st.button("Buscar mejores distribuciones", type="primary"):
        trimmed = trim_outliers(series, trim_percent)
        results, skipped_by_timeout = fit_many_distributions(trimmed, int(timeout_seconds))
        st.session_state.nonnormal_result = {
            "series": trimmed,
            "fits": results,
            "skipped_by_timeout": skipped_by_timeout,
            "il": il,
            "ul": ul,
        }
        st.session_state.nonnormal_applied_trim_percent = trim_percent
        st.session_state.selected_fit_name = results[0].name if results else None

    result = st.session_state.nonnormal_result
    if result:
        if not result["fits"]:
            st.error("No se pudo ajustar ninguna distribucion de forma confiable.")
            return

        skipped_by_timeout = result.get("skipped_by_timeout", [])
        if skipped_by_timeout:
            st.info(
                "Se saltaron por timeout: "
                + ", ".join(name.replace("_", " ").title() for name in skipped_by_timeout)
            )

        st.dataframe(results_to_dataframe(result["fits"]), use_container_width=True)
        good_fits = [item for item in result["fits"] if item.p_value > GOOD_FIT_P_VALUE]
        if good_fits:
            st.success(f"Hay {len(good_fits)} distribucion(es) con buen fit estadistico.")
        else:
            st.warning("Ninguna de las top 10 pasa el umbral de buen fit. Monte Carlo queda bloqueado para este modulo.")

        options = {item.display_name: item for item in result["fits"]}
        selected_display = st.selectbox("Distribucion a analizar", list(options.keys()))
        selected_fit = options[selected_display]
        distribution = getattr(stats, selected_fit.name)
        is_good_fit = selected_fit.p_value > GOOD_FIT_P_VALUE
        capability = nonnormal_capability(result["series"], distribution, selected_fit.params, result["il"], result["ul"])

        col1, col2, col3 = st.columns(3)
        col1.metric("Fit seleccionado", "Buen fit" if is_good_fit else "Fit debil")
        col2.metric("p-value KS", f"{selected_fit.p_value:.4f}")
        col3.metric("Ppk", f"{capability['ppk']:.3f}" if np.isfinite(capability["ppk"]) else "N/A")
        st.info(f"Estado del proceso: {capability['judgement']}")

        plot_distribution_analysis(
            result["series"],
            distribution,
            selected_fit.params,
            is_good_fit,
            result["il"],
            result["ul"],
            selected_fit.display_name,
        )


def monte_carlo_tab(series):
    st.subheader("Monte Carlo")
    normal_result = st.session_state.normal_result
    nonnormal_result = st.session_state.nonnormal_result

    sources = []
    if normal_result and normal_result["is_normal"]:
        sources.append("Normal")
    if nonnormal_result:
        good_fits = [fit for fit in nonnormal_result["fits"] if fit.p_value > GOOD_FIT_P_VALUE]
        if good_fits:
            sources.append("No normal")

    if not sources:
        st.warning("Monte Carlo se desbloquea cuando existe un buen fit normal o no normal.")
        return

    source = st.radio("Origen del fit", sources, horizontal=True)
    il, ul = specification_inputs("montecarlo")
    target_index = st.number_input("Cpk / Ppk deseado", min_value=0.1, value=1.33, step=0.01)
    n_simulations = st.number_input("Cantidad de simulaciones", min_value=1000, max_value=200000, value=20000, step=1000)

    if st.button("Simular Monte Carlo", type="primary"):
        if source == "Normal":
            fitted_series = normal_result["series"]
            distribution = stats.norm
            params = (fitted_series.mean(), fitted_series.std(ddof=1))
            label = "Normal"
        else:
            good_fits = [fit for fit in nonnormal_result["fits"] if fit.p_value > GOOD_FIT_P_VALUE]
            fit = good_fits[0]
            fitted_series = nonnormal_result["series"]
            distribution = getattr(stats, fit.name)
            params = fit.params
            label = fit.display_name

        st.session_state.monte_carlo_result = run_monte_carlo(
            label,
            distribution,
            params,
            fitted_series,
            il,
            ul,
            target_index,
            int(n_simulations),
        )

    if st.session_state.monte_carlo_result:
        st.write(st.session_state.monte_carlo_result)


def main():
    st.set_page_config(page_title="Process Capability Analysis", layout="wide")
    init_state()

    st.title("Process Capability Analysis")
    st.caption("Carga datos, evalua normalidad, calcula capacidad y simula escenarios objetivo.")

    uploaded_file = st.file_uploader(
        "Selecciona un archivo .csv o .xlsx",
        type=["csv", "xlsx", "xlsm", "xls"],
        help="Para .xlsx se exige una sola hoja, una sola tabla de Excel y ningun dato fuera de esa tabla.",
    )

    if uploaded_file and uploaded_file.name != st.session_state.file_name:
        try:
            df = load_uploaded_file(uploaded_file)
            if df.empty:
                raise ValueError("El archivo no contiene datos validos.")
            st.session_state.data = df
            st.session_state.file_name = uploaded_file.name
            clear_analysis_state()
            st.success("Archivo cargado y validado correctamente.")
        except Exception as exc:
            st.session_state.data = None
            st.session_state.file_name = None
            clear_analysis_state()
            st.error(str(exc))

    df = st.session_state.data
    if df is None:
        return

    st.dataframe(df, use_container_width=True)
    columns = numeric_columns(df)
    if not columns:
        st.error("El archivo no contiene columnas numericas analizables. No se aceptan fechas ni texto.")
        return

    selected_col = st.selectbox("Selecciona una columna numerica", columns)
    if selected_col != st.session_state.selected_col:
        st.session_state.selected_col = selected_col
        st.session_state.normal_result = None
        st.session_state.nonnormal_result = None
        st.session_state.monte_carlo_result = None

    series = clean_numeric_series(df, selected_col)
    if len(series) < 8:
        st.error("La columna debe tener al menos 8 valores numericos validos para analizar distribuciones.")
        return

    tab_normal, tab_nonnormal, tab_montecarlo = st.tabs(["Normal & Cpk", "No normal & Ppk", "Monte Carlo"])
    with tab_normal:
        normal_analysis_tab(series)
    with tab_nonnormal:
        nonnormal_analysis_tab(series)
    with tab_montecarlo:
        monte_carlo_tab(series)


if __name__ == "__main__":
    mp.freeze_support()
    main()
